# coding = utf8
import openpyxl
import sys
import os

base_path = os.path.pardir
sys.path.append(base_path)


# open_excel = openpyxl.load_workbook(base_path + r"\Case\Case_demo.xlsx")
# sheet_name = open_excel.sheetnames
# excel_value = open_excel[sheet_name[0]]
# print(excel_value)
# print(excel_value.cell(1, 1).value)


class HandExcel:
    def __init__(self, name):
        self.name = name
        pass

    def load_excel(self):
        # 加载excel  r"\Case\Case_demo.xlsx"
        try:
            open_excel = openpyxl.load_workbook(base_path + self.name)
            return open_excel
        except Exception as e:
            print(e)

    def get_sheet_data(self, index=None):
        # 记载所有的sheet
        sheet_name = self.load_excel().sheetnames
        if index is None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        # 获取某一个单元格内容
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        # 获取行数
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        # 获取某一行的内容
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list


# handle = HandExcel()

# if __name__ == "__main__":
#     handle = HandExcel()
#     print(handle.get_rows_value(1))
